import requests
from bs4 import BeautifulSoup
from urllib.parse import quote
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import urllib3

# Ignora warnings SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
requests.packages.urllib3.disable_warnings()

# =========================================================
# Função auxiliar para converter string de data em datetime
# =========================================================
def parse_data(data_str):
    try:
        return datetime.strptime(data_str, "%d/%m/%Y")
    except ValueError:
        return None

# =========================================================
# Lista final de eventos combinados
# =========================================================
eventos = []

# =========================================================
# 1️⃣ Coleta do site: brasilquecorre.com/santacatarina
# =========================================================
url_bqc = "https://brasilquecorre.com/santacatarina"
resp = requests.get(url_bqc, verify=False)
soup = BeautifulSoup(resp.text, "html.parser")

widgets = soup.find_all("div", class_="text-editor") + soup.find_all("article")

meses = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12
}

for w in widgets:
    titulo_tag = w.find("h5") or w.find("h2") or w.find("h3")
    link_tag = w.find("a", href=True)
    info_tags = w.find_all("p") or w.find_all("li")

    titulo = titulo_tag.get_text(strip=True).upper() if titulo_tag else "N/A"
    link_evento = link_tag["href"] if link_tag else "N/A"

    data_evento = None
    cidade = None
    distancia = None

    if info_tags:
        for linha in info_tags:
            texto = linha.get_text(strip=True)
            if not texto:
                continue

            # Procura data
            match = re.search(r'(\d{1,2}) de (\w+) de (\d{4})', texto)
            if match:
                dia, mes_txt, ano = match.groups()
                mes_num = meses.get(mes_txt.lower())
                if mes_num:
                    data_evento = f"{int(dia):02d}/{mes_num:02d}/{ano}"
            else:
                # Armazena cidade/distância se disponível
                if not cidade:
                    cidade = texto
                elif not distancia:
                    distancia = texto

    if data_evento:
        eventos.append({
            "titulo": titulo,
            "data": data_evento,
            "cidade": cidade or "N/A",
            "distancia": distancia or "N/A",
            "link": link_evento
        })

# =========================================================
# 2️⃣ Coleta do site: corridasbr.com.br/sc/Calendario
# =========================================================
base_url = "https://www.corridasbr.com.br/sc/Calendario"
extensao = ".asp"
pagina_num = 0

while True:
    url = f"{base_url}{extensao}" if pagina_num == 0 else f"{base_url}{pagina_num}{extensao}"
    resp = requests.get(url, verify=False)
    if resp.status_code != 200:
        break

    soup = BeautifulSoup(resp.text, "html.parser")
    trs = soup.find_all("tr", class_="tipo4")
    if not trs:
        break

    for w in trs:
        links = w.find_all("a", href=True)
        titulo = links[1].get_text(strip=True).upper() if len(links) > 1 else "N/A"

        # Monta link do evento
        if len(links) > 1:
            raw_link = links[1]["href"]
            segundo_link = quote(raw_link, safe="/:&?=")  # evita quebrar o link
        else:
            segundo_link = ""
        link_evento = "https://www.corridasbr.com.br/sc/" + segundo_link if segundo_link else "N/A"

        # Extrai data real da página do evento
        data_real = "N/A"
        if link_evento != "N/A":
            try:
                pagina_evento = requests.get(link_evento, verify=False)
                soup_evento = BeautifulSoup(pagina_evento.text, "html.parser")
                for td in soup_evento.find_all("td"):
                    texto_td = td.get_text(strip=True)
                    match = re.search(r'\d{2}/\d{2}/\d{4}', texto_td)
                    if match:
                        data_real = match.group()
                        break
            except requests.RequestException:
                pass

        info = w.find_all("td")
        cidade = info[1].get_text(strip=True) if len(info) > 1 else "N/A"
        distancia = info[3].get_text(strip=True) if len(info) > 3 else "N/A"

        if data_real != "N/A":
            eventos.append({
                "titulo": titulo,
                "data": data_real,
                "cidade": cidade,
                "distancia": distancia,
                "link": link_evento
            })

    pagina_num += 1

# =========================================================
# Remover duplicados (mesmo título + mesma data)
# =========================================================
eventos_unicos = []
vistos = set()
for ev in eventos:
    chave = (ev["titulo"], ev["data"])
    if chave not in vistos:
        vistos.add(chave)
        eventos_unicos.append(ev)

# =========================================================
# Ordenar por data crescente
# =========================================================
eventos_ordenados = sorted(eventos_unicos, key=lambda x: parse_data(x["data"]) or datetime.max)

# =========================================================
# Impressão no terminal
# =========================================================
for ev in eventos_ordenados:
    print("🏃 Corrida:", ev["titulo"])
    print("📌", ev["data"])
    print("📌", ev["cidade"])
    print("📌", ev["distancia"])
    print("🔗 Link:", ev["link"])
    print("-" * 60)

# =========================================================
# Exportar para Excel (corridas.xlsx)
# =========================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Corridas SC"

# Cabeçalho
headers = ["Data", "Cidade", "Título", "Distância", "Link"]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Linhas
for ev in eventos_ordenados:
    ws.append([ev["data"], ev["cidade"], ev["titulo"], ev["distancia"], ev["link"]])

# Centralizar todas as células
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Ajuste de largura automática (seguro)
for coluna in ws.columns:
    try:
        col_letter = coluna[0].column_letter
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in coluna)
        ws.column_dimensions[col_letter].width = max_length + 2
    except IndexError:
        continue

# Salvar arquivo
wb.save("corridas.xlsx")
print("\n✅ Arquivo 'corridas.xlsx' gerado com sucesso!")
